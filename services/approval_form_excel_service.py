"""
文件说明：
- 为审批单目录生成分组统计 Excel。

主要职责：
- 按文件名中“第一个下划线后到最后一个短横线前”的内容进行分组。
- 统计每组最后编码的最大值，并生成顺序编码列表。
- 导出多 sheet Excel，并自动设置列宽以完整显示字符串内容。

运行方式：
- 分类：被依赖脚本
- 直接运行命令：不建议直接运行
- 直接运行用途：无独立业务入口，主要被 workflow 调用。
- 被谁调用：workflows.approval_form_workflow
- 作为依赖用途：为审批单批处理流程输出统计台账。

输入：
- 配置输入：approval_form_prefix
- 数据输入：目标目录中的 PDF 文件列表
- 前置条件：已安装 openpyxl；PDF 文件名符合审批单命名规则

输出：
- 结果输出：Excel 文件路径
- 日志输出：调用方 logger
- 副作用：覆盖写出目标 Excel 文件

核心入口：
- 关键函数：export_approval_form_excel()

依赖关系：
- 依赖的本项目模块：无
- 依赖的第三方库：openpyxl
"""

from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import PatternFill
from openpyxl.styles import Side


INVALID_SHEET_CHARS_PATTERN = re.compile(r"[:\\/?*\[\]]")
MAX_SHEET_NAME_LENGTH = 31
GREEN_FILL = PatternFill(
    fill_type="solid",
    start_color="FF92D050",
    end_color="FF92D050",
)
YELLOW_FILL = PatternFill(
    fill_type="solid",
    start_color="FFFFFF99",
    end_color="FFFFFF99",
)
THIN_BORDER = Border(
    left=Side(style="thin", color="FFBFBFBF"),
    right=Side(style="thin", color="FFBFBFBF"),
    top=Side(style="thin", color="FFBFBFBF"),
    bottom=Side(style="thin", color="FFBFBFBF"),
)


def _display_width(text):
    width = 0
    for char in str(text):
        width += 2 if unicodedata.east_asian_width(char) in {"F", "W"} else 1
    return width


def _sanitize_sheet_name(sheet_name, used_names):
    sanitized = INVALID_SHEET_CHARS_PATTERN.sub("_", str(sheet_name).strip())
    sanitized = sanitized or "Sheet"
    sanitized = sanitized[:MAX_SHEET_NAME_LENGTH]

    candidate = sanitized
    counter = 2
    while candidate in used_names:
        suffix = f"_{counter}"
        base = sanitized[: MAX_SHEET_NAME_LENGTH - len(suffix)]
        candidate = f"{base}{suffix}"
        counter += 1

    used_names.add(candidate)
    return candidate


def _parse_pdf_name(pdf_path):
    stem = Path(pdf_path).stem
    first_underscore = stem.find("_")
    last_dash = stem.rfind("-")
    if first_underscore < 0 or last_dash < 0 or first_underscore >= last_dash:
        return None

    group_name = stem[first_underscore + 1 : last_dash].strip()
    code_text = stem[last_dash + 1 :].strip()
    if not group_name or not code_text.isdigit():
        return None

    return {
        "group_name": group_name,
        "code_value": int(code_text),
        "code_width": len(code_text),
    }


def _build_grouped_sequences(pdf_files, logger):
    grouped = {}
    invalid_names = []

    for pdf_path in sorted(Path(path) for path in pdf_files):
        parsed = _parse_pdf_name(pdf_path)
        if parsed is None:
            invalid_names.append(pdf_path.name)
            continue

        group_info = grouped.setdefault(
            parsed["group_name"],
            {"max_code": 0, "code_width": parsed["code_width"]},
        )
        group_info["max_code"] = max(group_info["max_code"], parsed["code_value"])
        group_info["code_width"] = max(group_info["code_width"], parsed["code_width"])

    if invalid_names:
        logger.warning(
            "以下 PDF 文件名不符合审批单统计规则，已跳过 Excel 统计: "
            + ", ".join(invalid_names)
        )

    return grouped


def export_approval_form_excel(
    pdf_files,
    logger,
    output_path,
    prefix="审批单_",
):
    grouped_sequences = _build_grouped_sequences(pdf_files, logger)
    if not grouped_sequences:
        logger.warning("没有可用于审批单统计的 PDF 文件，跳过 Excel 生成。")
        return False

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_sheet_names = set()
    existing_pdf_stems = set()
    for pdf_path in pdf_files:
        stem = Path(pdf_path).stem
        if prefix and stem.startswith(prefix):
            stem = stem[len(prefix) :]
        existing_pdf_stems.add(stem)

    for group_name in sorted(grouped_sequences):
        group_info = grouped_sequences[group_name]
        sheet_name = _sanitize_sheet_name(group_name, used_sheet_names)
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet["A1"] = "设计变更编号"
        worksheet["B1"] = "审批单情况"
        worksheet["A1"].border = THIN_BORDER
        worksheet["B1"].border = THIN_BORDER

        values = []
        for code_value in range(1, group_info["max_code"] + 1):
            code_text = str(code_value).zfill(group_info["code_width"])
            values.append(f"{group_name}-{code_text}")

        for row_index, value in enumerate(values, start=2):
            cell = worksheet.cell(row=row_index, column=1, value=value)
            cell.number_format = "@"
            cell.alignment = Alignment(horizontal="left")
            cell.border = THIN_BORDER

            status_value = "有审批单" if value in existing_pdf_stems else ""
            status_cell = worksheet.cell(row=row_index, column=2, value=status_value)
            status_cell.alignment = Alignment(horizontal="left")
            status_cell.border = THIN_BORDER

        max_a_width = max(_display_width(value) for value in values + ["设计变更编号"])
        max_b_width = max(_display_width(value) for value in ["审批单情况", "有审批单"])
        worksheet.column_dimensions["A"].width = max(max_a_width + 2, 12)
        worksheet.column_dimensions["B"].width = max(max_b_width + 2, 12)

        last_row = len(values) + 1
        target_range = f"A2:B{last_row}"
        worksheet.conditional_formatting.add(
            target_range,
            FormulaRule(
                formula=['$B2="有审批单"'],
                fill=GREEN_FILL,
                stopIfTrue=False,
            ),
        )
        worksheet.conditional_formatting.add(
            target_range,
            FormulaRule(
                formula=['LEN(TRIM($B2))=0'],
                fill=YELLOW_FILL,
                stopIfTrue=False,
            ),
        )

    excel_path = Path(output_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
    logger.info(f"审批单统计 Excel 已生成: {excel_path}")
    return excel_path
